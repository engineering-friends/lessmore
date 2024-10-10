from datetime import datetime

from lessmore.utils.system.open_in_os import open_in_os
from openpyxl import load_workbook


def generate_invoice(
    template_filename: str = "template.xlsx",
    replacements: dict = {},
    output_xlsx: str = "generated_file.xlsx",
):
    # - Load the Excel template

    workbook = load_workbook(template_filename)
    sheet = workbook.active  # Assuming you're working with the first sheet

    # - Replace placeholders in the Excel sheet

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, value in replacements.items():
                    if "{" + key + "}" not in str(cell.value):
                        continue
                    cell.value = cell.value.replace("{" + key + "}", str(value))

                    if key.startswith("INT_"):
                        cell.value = int(cell.value)

                if "{" in str(cell.value):
                    print(cell.value)
                    raise Exception("Failed to replace all placeholders")

    # - Save the modified Excel document

    workbook.save(output_xlsx)


if __name__ == "__main__":
    now = datetime.now()
    output_xlsx = "generated_file.xlsx"
    generate_invoice(
        template_filename="generate_invoice_template.xlsx",
        replacements={
            "N": 23,
            "TODAY_MM/YY": now.strftime("%m/%y"),
            "TODAY_DD.MM.YYYY": now.strftime("%d.%m.%Y"),
            "TODAY_YYYY-MM-DD": now.strftime("%Y-%m-%d"),
            "SERVICE_AGREEMENT": "Service Agreement No. 2024-001",
            "PAID_MONTH_YYYY-MM": "2024-09",
            "HOURS": 120,
            "INT_AMOUNT": 5000,
            "AMOUNT_WORDS": "Five Thousand",
        },
        output_xlsx=output_xlsx,
    )
    open_in_os(output_xlsx)
